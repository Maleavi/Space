import space
import math
import numpy as np
from openpyxl import Workbook

def array_reader(array_thing):
    #numpy doing shenenigans, so just made a generator for that
    for o in array_thing:
        yield o
    
#xlsx writer
wb = Workbook()
dest_filename = 'resources/data.xlsx'
ws1 = wb.active
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()

ws1.title = 'orbital positions'
ws2.title = 'orbital velocities'
ws3.title = 'gravity'

#parameters
pos = np.array([-1.0,0.0])
vel = np.array([0.0, 1.35])
actuness = 512
steps = 1024

grav = 0

for i in range(steps):
    pos = pos + (vel / actuness)
    vel = vel + (space.gravity(pos) / (actuness))
    grav = 0

    for d in space.gravity(pos):
        grav = grav + (d * d)

    grav = math.sqrt(grav)

    print (pos)
    print (vel)
    print (grav)
    print ('')
    ws1.append(array_reader(pos))
    ws2.append(array_reader(vel))
    ws3.append([grav])

wb.save(filename = dest_filename)
