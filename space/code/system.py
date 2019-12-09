import space
import math
import numpy as np
from openpyxl import Workbook

def array_reader(array_thing):
    #numpy doing shenenigans, so just made a generator for that
    for o in array_thing:
        yield o
    
#xlsx writer
wb = Workbook()
dest_filename = 'resources/data.xlsx'
ws1 = wb.active
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
ws4 = wb.create_sheet()
#ws5 = wb.create_sheet()
#ws6 = wb.create_sheet()

ws1.title = 'orbital position obj1'
ws2.title = 'orbital velocitie obj1'
ws3.title = 'orbital position obj2'
ws4.title = 'orbital velocitie obj2'
#ws5.title = 'orbital position obj3'
#ws6.title = 'orbital velocitie obj3'

#parameters
pos = np.array([-1.0,0.0])
vel = np.array([0.0, 0.5])
actuness = 512
steps = 8192

space_objects = []
space_objects.append(space.space_object(1, np.array([0.0, 0.0]), np.array([0.0, -0.5])))
#space_objects.append(space.space_object(1, np.array([0.0, -1.0]), np.array([0.0, -0.5])))
space_objects.append(space.space_object(1, pos, vel))


for i in range(steps):
    for obj in space_objects:
        obj.move(float(1.0/actuness))
    
    for obj in space_objects:
        forces = np.array([0.0, 0.0])
        for objs in space_objects:
            if not obj == objs:
                forces = forces + objs.grav(obj.pos)

        obj.accelerate(forces, float(1.0/actuness))

    print (space_objects[0].pos)
    print (space_objects[0].vel)
    print ('')
    print (space_objects[1].pos)
    print (space_objects[1].vel)
    print('')
    #print (space_objects[2].pos)
    #print (space_objects[2].vel)
    print('')
    ws1.append(array_reader(space_objects[0].pos))
    ws2.append(array_reader(space_objects[0].vel))
    ws3.append(array_reader(space_objects[1].pos))
    ws4.append(array_reader(space_objects[1].vel))
    #ws5.append(array_reader(space_objects[2].pos))
    #ws6.append(array_reader(space_objects[2].vel))

wb.save(filename = dest_filename)
